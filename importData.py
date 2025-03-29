import pandas as pd

# Importing Excel file
excel_df = pd.read_excel("Expenses.xlsx")
print("Excel Data:\n", excel_df.head())

# Importing CSV file
csv_df = pd.read_csv("Data.csv")
print("\nCSV Data:\n", csv_df.head())

# Importing JSON file
json_df = pd.read_json("data.json")
print("\nJSON Data:\n", json_df.head())

# Checking  missing values
print("\nMissing Values in Excel Data:\n", excel_df.isnull().sum())
print("\nMissing Values in CSV Data:\n", csv_df.isnull().sum())
print("\nMissing Values in JSON Data:\n", json_df.isnull().sum())

# Check  duplicates
print("\nDuplicate Rows in Excel Data:", excel_df.duplicated().sum())
print("Duplicate Rows in CSV Data:", csv_df.duplicated().sum())
print("Duplicate Rows in JSON Data:", json_df.duplicated().sum())

# Check data types
print("\nData Types in Excel Data:\n", excel_df.dtypes)
print("\nData Types in CSV Data:\n", csv_df.dtypes)
print("\nData Types in JSON Data:\n", json_df.dtypes)

# Get basic statistics
print("\nStatistics for Excel Data:\n", excel_df.describe())
print("\nStatistics for CSV Data:\n", csv_df.describe())
print("\nStatistics for JSON Data:\n", json_df.describe())

excel_df.dropna(inplace=True)
csv_df.dropna(inplace=True)
json_df.dropna(inplace=True)

excel_df.drop_duplicates(inplace=True)
csv_df.drop_duplicates(inplace=True)
json_df.drop_duplicates(inplace=True)

# ----Data Transformation & Analysis ----

# Filtering Data (Example: Select people older than 25)
filtered_csv = csv_df[csv_df["age"] > 25]
print("\nPeople older than 25:\n", filtered_csv)

# Grouping Data (Example: Count number of people per city)
grouped_csv = csv_df.groupby("city")["id"].count()
print("\nNumber of people per city:\n", grouped_csv)

# Sorting Data (Example: Sort expenses by cost)
sorted_excel = excel_df.sort_values(by="Item Cost", ascending=False)
print("\nSorted Expenses (Highest to Lowest):\n", sorted_excel.head())

# Applying Functions (Example: Calculate total expenses)
total_expenses = excel_df["Item Cost"].sum()
print("\nTotal Expenses: ", total_expenses)

# Save transformed data
filtered_csv.to_csv("filtered_people.csv", index=False)
grouped_csv.to_csv("people_per_city.csv")
sorted_excel.to_excel("sorted_expenses.xlsx", index=False)

import matplotlib.pyplot as plt
import seaborn as sns

# ----Data Visualization ----

# Set style for Seaborn
sns.set(style="whitegrid")

# 📊 1. Bar plot – Number of people per city
plt.figure(figsize=(8,5))
sns.barplot(x=grouped_csv.index, y=grouped_csv.values, palette="Blues_d")
plt.title("Number of People per City")
plt.xlabel("City")
plt.ylabel("Count")
plt.tight_layout()
plt.savefig("people_per_city_barplot.png")  # Save the plot as an image
plt.show()

# 📈 2. Line chart – Expenses over time (if "Date" exists)
if "Date" in excel_df.columns:
    plt.figure(figsize=(10,5))
    sns.lineplot(data=excel_df, x="Date", y="Item Cost", marker="o")
    plt.title("Expenses Over Time")
    plt.xlabel("Date")
    plt.ylabel("Cost")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig("expenses_over_time.png")
    plt.show()

import pandas as pd

# ----Export Processed Data ----

# Save cleaned & sorted Excel data
excel_df.to_excel("processed_expenses.xlsx", index=False)

# Save cleaned & grouped CSV data
csv_df.to_csv("processed_people.csv", index=False)

# Save JSON format of CSV data
csv_df.to_json("processed_people.json", orient="records")

print("✅ Data successfully exported!")

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference

# ----Generate an Improved Report ----

# Load data
df = pd.read_excel("processed_expenses.xlsx")

# Convert 'Date' column to readable format
df["Date"] = pd.to_datetime(df["Date"]).dt.strftime('%Y-%m-%d')

# Remove "TOTAL" row from the data
df_filtered = df[df["Expenses"].str.lower() != "total"]

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Expenses Report"

# Write headers with formatting
headers = list(df_filtered.columns)
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Write data to Excel
for r_idx, row in enumerate(df_filtered.values, start=2):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Adjust column width
for col in ws.columns:
    max_length = max(len(str(cell.value)) for cell in col if cell.value)
    ws.column_dimensions[col[0].column_letter].width = max_length + 3

# Add a total row at the bottom
total_expenses = df_filtered["Item Cost"].sum()
ws.append(["Total", total_expenses, "", total_expenses])
total_row = len(df_filtered) + 2
ws[f"A{total_row}"].font = Font(bold=True)
ws[f"B{total_row}"].number_format = 'KES #,##0'

# ---- Create a Cleaner Chart ----
chart = BarChart()
chart.title = "Expenses Breakdown"
chart.x_axis.title = "Expense Items"
chart.y_axis.title = "Cost (KES)"
chart.style = 12  # Modern style

# Select top 10 most expensive items for better visibility
df_top10 = df_filtered.nlargest(10, "Item Cost")

# Insert only the top 10 into the chart
data = Reference(ws, min_col=2, min_row=2, max_row=11, max_col=2)
categories = Reference(ws, min_col=1, min_row=2, max_row=11)

chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

# Resize and reposition chart
chart.width = 18
chart.height = 8
chart.legend.position = "b"  # Move legend below

ws.add_chart(chart, "E5")

# Save the improved report
wb.save("Improved_Expenses_Report.xlsx")

print("✅ Report successfully enhanced!")
